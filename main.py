"""Ponto de entrada — Automação Débitos em Aberto.

Fluxo:
    1. Abre janela para seleção da planilha (ui_upload)
    2. Lê a aba 'Empresas' da planilha:
         Coluna A = CNPJ
         Coluna B = EMPRESA
         Coluna C = CERTIFICADO
         Coluna D = resultado (preenchida pela automação)
    3. Ordena pela coluna C (certificado) para minimizar logins no eCAC
    4. Para o primeiro CNPJ de cada grupo de certificado:
         - Login no eCAC via LoginEcac
         - Navega para servicos.receitafederal.gov.br e autentica
    5. Para cada CNPJ (incluindo os subsequentes do mesmo certificado):
         - Aguarda intervalo mínimo de 30s entre trocas de CNPJ no portal
         - Representa o CNPJ como Procurador no portal
         - Navega para a página de pendências
         - Verifica status e escreve resultado na coluna D
"""

import argparse
import difflib
import json
import logging
import os
import openpyxl
from openpyxl.styles import Alignment
import re
import sys
import time
import unicodedata
from pathlib import Path

import pandas as pd
from dotenv import load_dotenv

# LoginEcac / captcha_uipath: suporta layout do repo (pacotes na mesma pasta)
# e layout de desenvolvimento legado (pastas irmãs LoginEcac / CaptchaSolver)
if getattr(sys, 'frozen', False):
    LOGIN_ECAC_DIR = Path(sys.executable).parent
else:
    _ecac_here    = Path(__file__).parent / "ecac_login"
    _ecac_sibling = Path(__file__).parent.parent / "LoginEcac"
    _cap_here     = Path(__file__).parent / "captcha_uipath"
    _cap_sibling  = Path(__file__).parent.parent / "CaptchaSolver"
    if not _ecac_here.exists() and _ecac_sibling.exists():
        sys.path.insert(0, str(_ecac_sibling))
        LOGIN_ECAC_DIR = _ecac_sibling
    else:
        LOGIN_ECAC_DIR = Path(__file__).parent
    if not _cap_here.exists() and _cap_sibling.exists():
        sys.path.insert(0, str(_cap_sibling))

from ecac_login import abrir_browser_com_certificado  # noqa: E402
from captcha_uipath import solve_hcaptcha                                  # noqa: E402
from ui_upload import main as selecionar_planilha                         # noqa: E402

CERTIFICADOS_DIR = Path(r"C:\Certificados")   # pode ser sobrescrito em main()
SENHAS_JSON      = CERTIFICADOS_DIR / "senhas.json"

# Chave Gemini — configure via .env (GEMINI_API_KEY=sua_chave) ou variável de ambiente
_GEMINI_API_KEY_PADRAO = os.environ.get("GEMINI_API_KEY", "")

URL_SERVICOS_RF        = "https://servicos.receitafederal.gov.br/"
URL_PENDENCIAS         = "https://servicos.receitafederal.gov.br/servico/pendencias/"
URL_ANALISE_PENDENCIAS = "https://servicos.receitafederal.gov.br/servico/pendencias/#/analise-pendencias"
# Classe específica do botão gov.br — mais robusto que XPath posicional
# (o botão contém <img alt="gov.br">, por isso has-text("gov.br") não funciona)
BTN_ENTRAR_GOV  = 'button.login-banner-button'

XPATH_STATUS = (
    "xpath=/html/body/app-root/mf-portal-layout/portal-main-layout/div/main/"
    "ng-component/app-consultar-dividas-pendencias/div[1]/app-resultado-analise-fiscal/"
    "div/div[2]/span"
)

# ── Timer de troca de CNPJ ────────────────────────────────────────────────────
# O portal não permite trocar de CNPJ com intervalo menor que 30 segundos.
_ultimo_troca_cnpj: float = 0.0
_INTERVALO_TROCA           = 30   # segundos


class FalhaPermanente(Exception):
    """Erro permanente que impede processar este CNPJ — não retentar.
    Exemplos: procuração expirada/inválida, CNPJ não autorizado pelo portal.
    """


# Palavras que o portal exibe em span.mensagemErro para indicar que o CNPJ
# não pode ser representado (procuração inexistente, vencida, CNPJ inválido…).
# Erros com essas palavras levantam FalhaPermanente em vez de serem retentados.
_PALAVRAS_ERRO_PERMANENTE = (
    "procuração", "procuracao", "vencid", "expirad",
    "não possui", "sem procuração", "não encontrad",
    "cnpj inválid", "não autorizado",
)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _remover_acentos(texto: str) -> str:
    """Remove acentos e diacríticos (NFD → ASCII)."""
    return unicodedata.normalize("NFD", texto).encode("ascii", "ignore").decode("ascii")


def _normalizar_cnpj(valor) -> str:
    """Remove qualquer formatação de CNPJ e retorna apenas 14 dígitos."""
    return re.sub(r"\D", "", str(valor)).zfill(14)


def _aguardar_intervalo_troca() -> None:
    """Aguarda o intervalo mínimo de 30s entre trocas de CNPJ no portal.

    Chamado sempre antes de clicar em 'Representar'. Se o intervalo já passou,
    retorna imediatamente sem bloqueio.
    """
    global _ultimo_troca_cnpj
    if _ultimo_troca_cnpj == 0.0:
        return
    decorrido = time.time() - _ultimo_troca_cnpj
    if decorrido < _INTERVALO_TROCA:
        espera = _INTERVALO_TROCA - decorrido
        print(f"    → Aguardando {espera:.0f}s (intervalo mínimo de {_INTERVALO_TROCA}s entre trocas)...")
        time.sleep(espera)


def _goto_seguro(page, url: str, label: str = "", timeout: int = 60_000) -> None:
    """Navega para `url` com logging detalhado e diagnóstico em caso de falha.

    Usa wait_until='domcontentloaded' em vez de 'networkidle':
    portais Angular mantêm conexões abertas e nunca atingem networkidle
    dentro de 30s, causando TimeoutError mesmo quando a página está pronta.
    O conteúdo real é verificado por wait_for() nos elementos seguintes.
    """
    prefixo = f"[{label}] " if label else ""
    url_antes = page.url
    print(f"    → {prefixo}Navegando para {url}\n"
          f"          (URL atual: {url_antes[:100]})")
    _t0 = time.time()
    try:
        page.goto(url, wait_until="domcontentloaded", timeout=timeout)
    except Exception as _err:
        _elapsed = time.time() - _t0
        _url_apos = page.url
        try:
            _titulo = page.title()
        except Exception:
            _titulo = "(indisponível)"
        print(
            f"    [!] {prefixo}Falha na navegação após {_elapsed:.1f}s:\n"
            f"          Erro       : {type(_err).__name__}: {_err}\n"
            f"          URL origem : {url_antes}\n"
            f"          URL destino: {url}\n"
            f"          URL atual  : {_url_apos}\n"
            f"          Título pág.: {_titulo!r}"
        )
        raise
    _elapsed = time.time() - _t0
    print(f"    [✓] {prefixo}Carregado em {_elapsed:.1f}s. URL: {page.url[:100]}")


def _aguardar_networkidle(page, timeout: int = 60_000, label: str = "") -> None:
    """Aguarda networkidle com fallback gracioso para SPAs Angular.

    SPAs Angular podem manter conexões abertas indefinidamente.
    Em vez de lançar exceção no timeout, registra o aviso e prossegue —
    o elemento-alvo é verificado pelo wait_for() da etapa seguinte.
    """
    try:
        page.wait_for_load_state("networkidle", timeout=timeout)
    except Exception as _err:
        _label_txt = f"[{label}] " if label else ""
        print(f"    → {_label_txt}networkidle não atingido "
              f"({type(_err).__name__}) — prosseguindo. URL: {page.url[:80]}")


# ── Certificados ──────────────────────────────────────────────────────────────

def _resolver_dir_certificados() -> Path:
    """Resolve a pasta de certificados na seguinte ordem:

    1. C:\\Certificados                    — caminho padrão (máquina do dev)
    2. <pasta do script>\\Certificados     — ao lado do main.py / iniciar.bat
    3. Dialog tkinter                      — usuário seleciona manualmente
    """
    import tkinter as tk
    from tkinter import filedialog, messagebox

    def _valida(p: Path) -> bool:
        return p.is_dir() and (p / "senhas.json").exists()

    # 1. Caminho padrão
    padrao = Path(r"C:\Certificados")
    if _valida(padrao):
        print(f"[cert] Pasta de certificados: {padrao}")
        return padrao

    # 2. Pasta 'Certificados' ao lado do executável (frozen) ou do script (dev)
    _base = Path(sys.executable).parent if getattr(sys, 'frozen', False) else Path(__file__).parent
    local = _base / "Certificados"
    if _valida(local):
        print(f"[cert] Pasta de certificados: {local}")
        return local

    # 3. Nenhum caminho padrão encontrado — pede ao usuário
    print("[!] Pasta de certificados não encontrada nos caminhos padrão.")
    print("    Abrindo seletor de pasta...")

    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)

    while True:
        pasta = filedialog.askdirectory(
            title="Selecione a pasta com os certificados (.pfx) e senhas.json",
            parent=root,
        )

        if not pasta:
            continuar = messagebox.askyesno(
                "Certificados não encontrados",
                "Nenhuma pasta selecionada.\nDeseja tentar novamente?",
                parent=root,
            )
            if not continuar:
                root.destroy()
                print("[!] Operação cancelada pelo usuário. Encerrando.")
                sys.exit(1)
            continue

        caminho = Path(pasta)
        if not (caminho / "senhas.json").exists():
            messagebox.showwarning(
                "Pasta inválida",
                f"O arquivo 'senhas.json' não foi encontrado em:\n{pasta}\n\n"
                "Selecione a pasta correta.",
                parent=root,
            )
            continue

        root.destroy()
        print(f"[cert] Pasta de certificados informada pelo usuário: {caminho}")
        return caminho


def carregar_certificados() -> dict[str, tuple[Path, str]]:
    """Lê senhas.json e retorna {nome_normalizado: (pfx_path, senha)}."""
    with open(SENHAS_JSON, encoding="utf-8") as f:
        senhas: dict[str, str] = json.load(f)
    return {
        _remover_acentos(Path(nome).stem.lower()): (CERTIFICADOS_DIR / nome, senha)
        for nome, senha in senhas.items()
    }


def _buscar_certificado(nome: str, certs: dict) -> str | None:
    """Retorna a chave normalizada de `certs` que corresponde ao nome do certificado."""
    chave = _remover_acentos(Path(nome).stem.lower())
    if chave in certs:
        return chave
    matches = difflib.get_close_matches(chave, certs.keys(), n=1, cutoff=0.75)
    if matches:
        print(f"    → Certificado '{nome}' resolvido para '{matches[0]}' (correspondência aproximada).")
        return matches[0]
    return None


def atualizar_env_certificado(pfx_path: Path, passphrase: str) -> None:
    """Atualiza CERT_PFX_PATH e CERT_PFX_PASSPHRASE no .env do LoginEcac."""
    env_path = LOGIN_ECAC_DIR / ".env"
    existentes: dict[str, str] = {}
    if env_path.exists():
        for linha in env_path.read_text(encoding="utf-8").splitlines():
            if "=" in linha and not linha.startswith("#"):
                chave, _, valor = linha.partition("=")
                existentes[chave.strip()] = valor.strip()
    if "GEMINI_API_KEY" not in existentes:
        existentes["GEMINI_API_KEY"] = os.environ.get("GEMINI_API_KEY", _GEMINI_API_KEY_PADRAO)
    existentes["CERT_PFX_PATH"]       = str(pfx_path)
    existentes["CERT_PFX_PASSPHRASE"] = passphrase
    conteudo = "\n".join(f"{k}={v}" for k, v in existentes.items()) + "\n"
    env_path.write_text(conteudo, encoding="utf-8")
    print(f"    [cert] Configurado: {pfx_path.name}")


# ── Planilha ──────────────────────────────────────────────────────────────────

def ler_e_ordenar(caminho: str) -> pd.DataFrame:
    """Lê a aba 'Empresas', remove duplicatas e ordena pela coluna C (certificado)."""
    ext = Path(caminho).suffix.lower()
    if ext in (".xlsx", ".xls"):
        df = pd.read_excel(caminho, sheet_name="Empresas", dtype=str)
    else:
        df = pd.read_csv(caminho, dtype=str)

    df = df.dropna(how="all").reset_index(drop=True)

    total_antes = len(df)
    df = df.drop_duplicates(ignore_index=True)
    removidas = total_antes - len(df)
    if removidas:
        print(f"  ⚠ {removidas} linha(s) duplicada(s) removida(s) da planilha.")

    col_certificado = df.columns[2]   # Coluna C = CERTIFICADO
    df = df.sort_values(by=col_certificado, kind="stable", ignore_index=True)
    return df


def escrever_coluna_d(caminho_planilha: str, cnpj: str, valor: str) -> None:
    """Abre a planilha com openpyxl e escreve `valor` na coluna D da linha
    cuja coluna A corresponde ao CNPJ informado (comparação normalizada)."""
    wb = openpyxl.load_workbook(caminho_planilha)
    ws = wb["Empresas"]

    for linha in ws.iter_rows(min_row=2):
        celula_cnpj = linha[0]   # Coluna A
        if celula_cnpj.value and _normalizar_cnpj(celula_cnpj.value) == cnpj:
            celula_d = linha[3]   # Coluna D
            celula_d.value = valor
            celula_d.alignment = Alignment(horizontal="center", vertical="center")
            break
    else:
        print(f"    [!] CNPJ {cnpj} não encontrado na planilha para escrita.")

    wb.save(caminho_planilha)
    print(f"    [✓] Coluna D → '{valor}'  (CNPJ {cnpj})")


def escrever_coluna_e(caminho_planilha: str, cnpj: str, valor: str) -> None:
    """Abre a planilha com openpyxl e escreve `valor` na coluna E da linha
    cuja coluna A corresponde ao CNPJ informado (comparação normalizada).

    Coluna E da aba 'Empresas' registra o status dos Processos Fiscais.
    """
    wb = openpyxl.load_workbook(caminho_planilha)
    ws = wb["Empresas"]

    for linha in ws.iter_rows(min_row=2):
        celula_cnpj = linha[0]   # Coluna A
        if celula_cnpj.value and _normalizar_cnpj(celula_cnpj.value) == cnpj:
            celula_e = linha[4]   # Coluna E
            celula_e.value = valor
            celula_e.alignment = Alignment(horizontal="center", vertical="center")
            break
    else:
        print(f"    [!] CNPJ {cnpj} não encontrado na planilha para escrita em E.")

    wb.save(caminho_planilha)
    print(f"    [✓] Coluna E → '{valor}'  (CNPJ {cnpj})")


def ler_status_cnpj(caminho_planilha: str, cnpj: str) -> tuple[str, str]:
    """Lê os valores das colunas D e E da aba 'Empresas' para o CNPJ dado.

    Retorna (val_d, val_e) — strings vazias quando as células estiverem em branco.
    Usado para verificar se o CNPJ já foi (parcialmente) processado.
    """
    wb = openpyxl.load_workbook(caminho_planilha)
    ws = wb["Empresas"]
    for linha in ws.iter_rows(min_row=2):
        celula_cnpj = linha[0]   # Coluna A
        if celula_cnpj.value and _normalizar_cnpj(celula_cnpj.value) == cnpj:
            val_d = str(linha[3].value or "").strip()
            val_e = str(linha[4].value or "").strip() if len(linha) > 4 else ""
            return val_d, val_e
    return "", ""


def escrever_aba_debitos(caminho_planilha: str, dados: list[dict]) -> None:
    """Adiciona os dados extraídos da tabela DCTFWeb na aba 'Débitos' da planilha.

    Se a aba ainda não existir, ela é criada com cabeçalho.
    Os dados são sempre acrescentados ao fim da aba (append).
    """
    wb = openpyxl.load_workbook(caminho_planilha)

    if "Débitos" not in wb.sheetnames:
        ws = wb.create_sheet("Débitos")
        cabecalho = [
            "CNPJ", "TIPO", "TRIBUTO", "Rec.", "PA/Ex.",
            "Dt.Vcto.", "Valor Original", "Saldo Devedor",
        ]
        ws.append(cabecalho)
    else:
        ws = wb["Débitos"]

    for d in dados:
        ws.append([
            d.get("cnpj", ""),
            d.get("tipo", ""),
            d.get("tributo", ""),
            d.get("receita", ""),
            d.get("pa_ex", ""),
            d.get("dt_vcto", ""),
            d.get("valor_original", ""),
            d.get("saldo", ""),
        ])

    wb.save(caminho_planilha)
    print(f"    [✓] Aba 'Débitos': {len(dados)} linha(s) gravada(s).")


# ── DCTFWeb: extração da tabela ────────────────────────────────────────────────

def _selecionar_n_por_pagina(page, n: int) -> None:
    """Seleciona N itens por página no ng-select de paginação (div.pagination-per-page).

    O ng-dropdown-panel é renderizado fora do container (appendTo body), por isso a
    opção é buscada diretamente pela span.ng-option-label com texto exato.
    """
    try:
        ng_sel = page.locator('div.pagination-per-page ng-select').first
        ng_sel.wait_for(state="visible", timeout=10_000)
        ng_sel.click()
        page.wait_for_timeout(600)

        opcao = page.locator(f'span.ng-option-label:text-is("{n}")').first
        opcao.wait_for(state="visible", timeout=5_000)
        opcao.click()
        page.wait_for_timeout(1_500)
        print(f"    [✓] Itens por página: {n}.")
    except Exception as exc:
        print(f"    [!] Não foi possível mudar para {n} itens/página: {exc}")


def _expandir_todas_as_linhas(page) -> None:
    """Expande todas as linhas da tabela de uma só vez via JavaScript.

    Um único evaluate clica todos os botões chevron-down simultaneamente,
    eliminando os N roundtrips Playwright + N × 350 ms da abordagem anterior.
    Aguarda 1 s para o Angular processar todas as mudanças de estado.
    """
    contagem = page.evaluate(
        """
        () => {
            const btns = Array.from(
                document.querySelectorAll('button.br-button.circle.small')
            ).filter(b => b.querySelector('i.fa-chevron-down'));
            btns.forEach(b => b.click());
            return btns.length;
        }
        """
    )
    if contagem:
        page.wait_for_timeout(1_000)   # aguarda Angular renderizar tudo
        print(f"    → {contagem} linha(s) expandida(s).")
    else:
        print("    → Nenhuma linha para expandir.")


def _extrair_dados_pagina(page, cnpj: str) -> list[dict]:
    """Extrai as linhas de dados visíveis na tabela DCTFWeb (JavaScript evaluate)."""
    return page.evaluate(
        """
        (cnpj) => {
            const resultado = [];

            /* Linhas principais: <tr> que contêm o botão de expandir */
            const linhas = document.querySelectorAll(
                'tbody tr:has(button.br-button.circle.small)'
            );

            for (const tr of linhas) {
                const tds = Array.from(tr.querySelectorAll('td'));

                /* Receita: única <td class="text-nowrap"> */
                const tdReceita = tds.find(td => td.classList.contains('text-nowrap'));
                const receita   = tdReceita?.textContent?.trim() ?? '';

                /* Saldo devedor consolidado: última <td class="text-right"> */
                const tdsSaldoDir = tds.filter(td => td.classList.contains('text-right'));
                const saldo = tdsSaldoDir[tdsSaldoDir.length - 1]
                                ?.textContent?.trim() ?? '';

                /* Tipo (Situação do débito): <td> que contém <span class="text-nowrap"> */
                const tdTipo = tds.find(td => td.querySelector('span.text-nowrap'));
                const tipo   = tdTipo?.querySelector('span.text-nowrap')
                                      ?.textContent?.trim() ?? '';

                /* <td> sem classe especial, sem botão e sem input (checkbox)
                   Ordem esperada no DOM: PA/Ex., Dt.Vcto., Saldo devedor (R$) */
                const tdsSimples = tds.filter(td =>
                    !td.classList.contains('text-nowrap') &&
                    !td.classList.contains('text-right') &&
                    !td.querySelector('span.text-nowrap') &&
                    !td.querySelector('button.br-button') &&
                    !td.querySelector('input')
                );
                const pa_ex   = tdsSimples[0]?.textContent?.trim() ?? '';
                const dt_vcto = tdsSimples[1]?.textContent?.trim() ?? '';

                /* Detalhe expandido: próxima <tr> irmã, revelada ao clicar na seta */
                let tributo       = '';
                let valor_original = '';
                const proxTr = tr.nextElementSibling;
                if (proxTr && proxTr.tagName === 'TR') {
                    const labels = proxTr.querySelectorAll('p.label');
                    for (const labelEl of labels) {
                        const labelText = labelEl.textContent.trim();
                        const divPai    = labelEl.closest('div');
                        const ps        = divPai
                            ? Array.from(divPai.querySelectorAll('p'))
                            : [];
                        /* Valor é o primeiro <p> que não tem class="label" */
                        const valorEl = ps.find(p => !p.classList.contains('label'));
                        const valor   = valorEl?.textContent?.trim() ?? '';
                        if (labelText === 'Tributo')              tributo        = valor;
                        if (labelText === 'Valor original (R$)')  valor_original = valor;
                    }
                }

                resultado.push({
                    cnpj, tipo, tributo, receita,
                    pa_ex, dt_vcto, valor_original, saldo,
                });
            }

            return resultado;
        }
        """,
        cnpj,
    )


def extrair_debitos_dctfweb(page, cnpj: str, caminho_planilha: str) -> None:
    """Após clicar no botão 'Dívida DCTFWeb', extrai todos os dados da tabela
    (paginando se necessário) e grava na aba 'Débitos' + coluna D (Empresas)."""

    print("    → Aguardando tabela DCTFWeb carregar...")
    _aguardar_networkidle(page, label="DCTFWeb")
    page.wait_for_timeout(1_000)

    # Muda para 50 itens por página
    _selecionar_n_por_pagina(page, 50)

    todos_dados: list[dict] = []
    pagina = 1

    while True:
        print(f"    → Extraindo dados (página {pagina})...")
        page.wait_for_timeout(1_500)

        _expandir_todas_as_linhas(page)

        dados_pag = _extrair_dados_pagina(page, cnpj)
        todos_dados.extend(dados_pag)
        print(f"    → {len(dados_pag)} linha(s) extraída(s) na página {pagina}.")

        # Verifica se há próxima página habilitada
        try:
            btn_proxima = page.locator(
                'button[aria-label="Página seguinte"]'
            ).first
            if not btn_proxima.is_disabled():
                print("    → Indo para próxima página...")
                btn_proxima.click()
                _aguardar_networkidle(page, label="DCTFWeb pág.")
                pagina += 1
                continue
        except Exception:
            pass
        break

    print(f"    [✓] Total: {len(todos_dados)} linha(s) de débito DCTFWeb.")
    escrever_aba_debitos(caminho_planilha, todos_dados)
    escrever_coluna_d(caminho_planilha, cnpj, "Concluído")


# ── Processo Fiscal: extração de cards ────────────────────────────────────────

def escrever_aba_processos_fiscais(caminho_planilha: str, dados: list[dict]) -> None:
    """Adiciona linhas na aba 'Processos Fiscais' (cria se não existir)."""
    wb = openpyxl.load_workbook(caminho_planilha)

    if "Processos Fiscais" not in wb.sheetnames:
        ws = wb.create_sheet("Processos Fiscais")
        ws.append([
            "CNPJ", "TIPO", "RECEITA", "PA/Ex.", "Dt.Vcto.",
            "Valor Original", "Saldo Devedor", "Processo de Crédito",
        ])
    else:
        ws = wb["Processos Fiscais"]

    for d in dados:
        ws.append([
            d.get("cnpj", ""),
            d.get("tipo", ""),
            d.get("receita", ""),
            d.get("pa_ex", ""),
            d.get("dt_vcto", ""),
            d.get("valor_original", ""),
            d.get("saldo", ""),
            d.get("processo_credito", ""),
        ])

    wb.save(caminho_planilha)
    print(f"    [✓] Aba 'Processos Fiscais': {len(dados)} linha(s) gravada(s).")


def _extrair_dados_pagina_processo(page, cnpj: str,
                                   processo_credito: str) -> list[dict]:
    """Extrai linhas da tabela de débitos de um card de processo fiscal."""
    return page.evaluate(
        """
        ([cnpj, processo_credito]) => {
            const resultado = [];
            const linhas = document.querySelectorAll(
                'tbody tr:has(button.br-button.circle.small)'
            );

            for (const tr of linhas) {
                const tds = Array.from(tr.querySelectorAll('td'));

                /* Receita: td.text-nowrap */
                const tdReceita = tds.find(td => td.classList.contains('text-nowrap'));
                const receita   = tdReceita?.textContent?.trim() ?? '';

                /* Saldo devedor: td.text-right */
                const tdSaldo = tds.find(td => td.classList.contains('text-right'));
                const saldo   = tdSaldo?.textContent?.trim() ?? '';

                /* Tipo: td contendo span.text-nowrap */
                const tdTipo = tds.find(td => td.querySelector('span.text-nowrap'));
                const tipo   = tdTipo?.querySelector('span.text-nowrap')
                                      ?.textContent?.trim() ?? '';

                /* TDs simples (sem classe, sem botão, sem input) → PA/Ex., Dt.Vcto. */
                const tdsSimples = tds.filter(td =>
                    !td.classList.contains('text-nowrap') &&
                    !td.classList.contains('text-right') &&
                    !td.querySelector('span.text-nowrap') &&
                    !td.querySelector('button.br-button') &&
                    !td.querySelector('input')
                );
                const pa_ex   = tdsSimples[0]?.textContent?.trim() ?? '';
                const dt_vcto = tdsSimples[1]?.textContent?.trim() ?? '';

                /* Valor original: seção expandida (próxima <tr> irmã) */
                let valor_original = '';
                const proxTr = tr.nextElementSibling;
                if (proxTr && proxTr.tagName === 'TR') {
                    const labels = proxTr.querySelectorAll('p.label');
                    for (const labelEl of labels) {
                        const labelText = labelEl.textContent.trim();
                        const divPai    = labelEl.closest('div');
                        const ps        = divPai
                            ? Array.from(divPai.querySelectorAll('p'))
                            : [];
                        const valorEl = ps.find(p => !p.classList.contains('label'));
                        const valor   = valorEl?.textContent?.trim() ?? '';
                        if (labelText === 'Valor original (R$)') valor_original = valor;
                    }
                }

                resultado.push({
                    cnpj, tipo, receita, pa_ex, dt_vcto,
                    valor_original, saldo, processo_credito,
                });
            }

            return resultado;
        }
        """,
        [cnpj, processo_credito],
    )


def _processar_card_processo(page, cnpj: str) -> list[dict]:
    """Extrai dados de um card de processo fiscal já aberto (nova página).

    Retorna lista de dicts com as linhas da tabela de débitos do card.
    """
    _aguardar_networkidle(page, label="card")
    page.wait_for_timeout(1_000)

    # ── "Processo de crédito" (expande se existir) ────────────────────────────
    processo_credito = ""
    btn_cred = page.locator(
        'button[aria-label="Expandir processo de crédito"]'
    ).first
    try:
        if btn_cred.is_visible(timeout=3_000):
            btn_cred.click()
            page.wait_for_timeout(800)
            # Após o clique o Angular muda o aria-label do botão (Expandir → Recolher),
            # por isso NÃO buscamos o botão novamente no JS.
            # Buscamos diretamente o div revelado: div.processo-credito
            try:
                div_proc = page.locator('div.processo-credito').first
                div_proc.wait_for(state="visible", timeout=3_000)
                processo_credito = (div_proc.text_content() or "").strip()
            except Exception:
                # Fallback via JS — cobre o caso de o div já existir mas sem estado "visible"
                processo_credito = page.evaluate(
                    "() => {"
                    "  const d = document.querySelector('div.processo-credito');"
                    "  return d ? d.textContent.trim() : '';"
                    "}"
                )
            print(f"    → Processo de crédito: {processo_credito}")
    except Exception:
        pass

    # ── Tabela de débitos do card ─────────────────────────────────────────────
    _selecionar_n_por_pagina(page, 50)

    todos_dados: list[dict] = []
    pagina = 1

    while True:
        print(f"    → Extraindo débitos do card (página {pagina})...")
        page.wait_for_timeout(1_500)
        _expandir_todas_as_linhas(page)

        dados_pag = _extrair_dados_pagina_processo(page, cnpj, processo_credito)
        todos_dados.extend(dados_pag)
        print(f"    → {len(dados_pag)} linha(s) na página {pagina}.")

        try:
            btn_prox = page.locator('button[aria-label="Página seguinte"]').first
            if not btn_prox.is_disabled():
                btn_prox.click()
                _aguardar_networkidle(page, label="card pág.")
                pagina += 1
                continue
        except Exception:
            pass
        break

    return todos_dados


def extrair_processo_fiscal(page, cnpj: str, caminho_planilha: str) -> None:
    """Percorre todos os cards de processo fiscal, extrai os dados de cada um
    e grava na aba 'Processos Fiscais' da planilha."""

    print("    → Aguardando página de processos fiscais carregar...")
    _aguardar_networkidle(page, label="proc.fiscal")
    page.wait_for_timeout(1_000)

    # Mostra 20 cards por página
    _selecionar_n_por_pagina(page, 20)

    todos_dados: list[dict] = []
    pagina_cards = 1
    SELETOR_CARD = (
        'button[aria-label="Expandir informações complementares do processo fiscal"]'
    )

    while True:
        print(f"    → Processando cards (página {pagina_cards})...")
        page.wait_for_timeout(1_500)

        total_cards = page.locator(SELETOR_CARD).count()
        print(f"    → {total_cards} card(s) nesta página.")

        for i in range(total_cards):
            print(f"    → Entrando no card {i + 1}/{total_cards}...")
            # Rebusca o botão fresco a cada iteração (evita stale reference)
            btn_card = page.locator(SELETOR_CARD).nth(i)
            btn_card.scroll_into_view_if_needed()
            btn_card.click()

            dados_card = _processar_card_processo(page, cnpj)
            todos_dados.extend(dados_card)

            # Volta para a lista de cards
            page.go_back()
            _aguardar_networkidle(page, label="voltar")
            page.wait_for_timeout(1_000)

        # Verifica próxima página de cards
        try:
            btn_prox = page.locator('button[aria-label="Página seguinte"]').first
            if not btn_prox.is_disabled():
                print("    → Indo para próxima página de cards...")
                btn_prox.click()
                _aguardar_networkidle(page, label="cards pág.")
                pagina_cards += 1
                continue
        except Exception:
            pass
        break

    print(f"    [✓] Total: {len(todos_dados)} linha(s) de processo fiscal.")
    escrever_aba_processos_fiscais(caminho_planilha, todos_dados)
    escrever_coluna_e(caminho_planilha, cnpj, "Concluído")


# ── Navegação — Portal ────────────────────────────────────────────────────────

def _fazer_logout(page) -> None:
    """Encerra a sessão no portal servicos.receitafederal.gov.br.

    Sequência:
        1. Clica no avatar (#avatar-dropdown-trigger) para abrir o menu.
        2. Clica em 'Sair' (#btn-sair).
        3. Confirma clicando no botão primário 'Sair' (button.br-button.is-primary).
    """
    print("    → Fazendo logout do portal...")
    try:
        avatar = page.locator('#avatar-dropdown-trigger').first
        avatar.wait_for(state="visible", timeout=8_000)
        avatar.click()
        page.wait_for_timeout(600)

        btn_sair = page.locator('#btn-sair').first
        btn_sair.wait_for(state="visible", timeout=5_000)
        btn_sair.click()
        page.wait_for_timeout(600)

        btn_confirmar = page.locator('button.br-button.is-primary').first
        btn_confirmar.wait_for(state="visible", timeout=5_000)
        btn_confirmar.click()
        page.wait_for_timeout(1_500)

        print("    [✓] Logout realizado com sucesso.")
    except Exception as e:
        print(f"    [!] Logout não foi possível (página talvez já encerrada): {e}")


def _fechar_navegador(p, context, page=None) -> None:
    """Faz logout (se `page` fornecida), fecha o contexto e para o Playwright."""
    if page is not None:
        _fazer_logout(page)
    try:
        context.close()
    except Exception:
        pass
    try:
        p.stop()
    except Exception:
        pass
    print("    [✓] Navegador fechado.")


def ir_para_servicos_rf_e_entrar(page) -> None:
    """Navega para servicos.receitafederal.gov.br, fecha popups e clica em
    'Entrar com gov.br'. Resolve hCaptcha se aparecer.

    Se aparecer 'Erro na validação do HCaptcha', volta para a página inicial do
    portal e repete o fluxo (até 3 tentativas no total).
    """
    MAX_TENTATIVAS_GOVBR = 3
    _fazer_go_back = False  # True quando o erro de captcha já acionou go_back()

    for tentativa_govbr in range(1, MAX_TENTATIVAS_GOVBR + 1):
        if tentativa_govbr > 1:
            print(f"    → Reiniciando fluxo 'Entrar com gov.br' "
                  f"(tentativa {tentativa_govbr}/{MAX_TENTATIVAS_GOVBR})...")

        if _fazer_go_back:
            _fazer_go_back = False
            # Já voltamos via go_back() — apenas aguarda a página estabilizar
            page.wait_for_load_state("domcontentloaded", timeout=15_000)
        else:
            print("    → Navegando para servicos.receitafederal.gov.br...")
            page.goto(URL_SERVICOS_RF, wait_until="domcontentloaded", timeout=30_000)

        # ── Popup 1: tourModal ────────────────────────────────────────────────
        try:
            popup1 = page.locator('xpath=//*[@id="tourModal"]/div[3]/a').first
            popup1.wait_for(state="visible", timeout=2_000)
            popup1.click()
            print("    [✓] Popup tourModal fechado.")
        except Exception:
            pass

        # ── Popup 2: card0 ───────────────────────────────────────────────────
        try:
            popup2 = page.locator('xpath=//*[@id="card0"]/div/div[2]/button[2]').first
            popup2.wait_for(state="visible", timeout=2_000)
            popup2.click()
            print("    [✓] Popup card0 fechado.")
        except Exception:
            pass

        # Clica "Entrar com gov.br" assim que o botão estiver visível.
        print("    → Aguardando botão 'Entrar com gov.br'...")
        _entrar_ok = False
        for _sel_entrar in [
            'xpath=//*[@id="home-heading"]/div[1]/div/button',
            BTN_ENTRAR_GOV,
        ]:
            try:
                _btn = page.locator(_sel_entrar).first
                _btn.wait_for(state="visible", timeout=20_000 if "xpath" in _sel_entrar else 3_000)
                _btn.click()
                print(f"    → Clicou 'Entrar com gov.br' ({_sel_entrar}).")
                _entrar_ok = True
                break
            except Exception:
                continue
        if not _entrar_ok:
            raise Exception("Botão 'Entrar com gov.br' não encontrado na página.")

        # ── Aguarda sso.acesso.gov.br e clica "Seu certificado digital" ──────
        # Após redirecionar para o gov.br (sso.acesso.gov.br), o botão
        # #login-certificate deve aparecer. Captcha pode preceder o botão.
        print("    → Aguardando 'Seu certificado digital' (sso.acesso.gov.br)...")
        _cert_deadline = time.time() + 40
        _cert_clicado  = False
        while time.time() < _cert_deadline and not _cert_clicado:
            # Tenta clicar no botão "Seu certificado digital"
            for _sel_cert in [
                'xpath=//*[@id="login-certificate"]',
                '#login-certificate',
                "a:has-text('Seu certificado digital')",
                "button:has-text('Seu certificado digital')",
            ]:
                try:
                    _loc = page.locator(_sel_cert).first
                    if _loc.is_visible(timeout=400):
                        _loc.click()
                        print(f"    → Clicou 'Seu certificado digital' ({_sel_cert}).")
                        _cert_clicado = True
                        break
                except Exception:
                    pass
            if _cert_clicado:
                # Aguarda até 10s por qualquer sinal de progresso após clicar cert:
                #   1. avatar apareceu  → já autenticado
                #   2. URL mudou        → redirecionamento em andamento
                #   3. iframe hcaptcha (checkbox OU challenge) visível → captcha ativo
                # Se nenhum dos três ocorrer em 10s, a página travou → recarrega.
                _url_antes_cert = page.url
                _cert_progrediu = False
                for _ in range(20):   # 20 × 500 ms = 10 s
                    page.wait_for_timeout(500)
                    try:
                        if page.locator('xpath=//*[@id="avatar-dropdown-trigger"]').is_visible():
                            _cert_progrediu = True
                            break
                    except Exception:
                        pass
                    if page.url != _url_antes_cert:
                        _cert_progrediu = True
                        break
                    try:
                        _hc_visivel = (
                            page.locator(
                                "iframe[src*='hcaptcha.com'][src*='frame=checkbox']"
                            ).is_visible()
                            or page.locator(
                                "iframe[src*='hcaptcha.com'][src*='frame=challenge']"
                            ).is_visible()
                        )
                        if _hc_visivel:
                            _cert_progrediu = True
                            break
                    except Exception:
                        pass
                if _cert_progrediu:
                    break  # prossegue para o poll de avatar
                print("    → Página travou após cert. Recarregando e retentando...")
                try:
                    page.reload(wait_until="domcontentloaded", timeout=15_000)
                except Exception:
                    pass
                _cert_clicado = False  # continua o while para tentar clicar novamente

            # Se avatar já apareceu, autenticação foi concluída sem precisar clicar cert
            try:
                if page.locator('xpath=//*[@id="avatar-dropdown-trigger"]').is_visible():
                    print("    → Avatar detectado antes do cert — já autenticado.")
                    _cert_clicado = True
                    break
            except Exception:
                pass

            # Resolve captcha se aparecer enquanto aguarda o botão cert
            _hc_frames_cert = [
                f for f in page.frames
                if "hcaptcha.com" in (f.url or "") and "frame=challenge" in (f.url or "")
            ]
            _captcha_cert = False
            for _hf in _hc_frames_cert:
                try:
                    _txt = _hf.evaluate("""() => {
                        const c = document.querySelector('.challenge-container');
                        if (!c) return null;
                        const r = c.getBoundingClientRect();
                        if (r.width < 100 || r.height < 100) return null;
                        const p = document.querySelector('.prompt-text');
                        if (!p || !p.textContent.trim()) return null;
                        const b = document.querySelector('.button-submit');
                        if (!b || b.getAttribute('aria-disabled') === 'true') return null;
                        return p.textContent.trim();
                    }""")
                    if _txt:
                        _captcha_cert = True
                        print(f"    → Captcha detectado antes do cert: '{_txt}'. Resolvendo...")
                        break
                except Exception:
                    pass
            if not _captcha_cert:
                try:
                    _captcha_cert = page.locator(
                        "iframe[src*='hcaptcha.com'][src*='frame=checkbox']"
                    ).is_visible()
                    if _captcha_cert:
                        print("    → Checkbox hCaptcha antes do cert. Resolvendo...")
                except Exception:
                    pass
            if _captcha_cert:
                _pre_cert_ok = False
                for _t in range(1, 3):   # máx. 2 tentativas
                    try:
                        if solve_hcaptcha(page):
                            print(f"    [✓] Captcha (pré-cert) resolvido ({_t}/2).")
                            _pre_cert_ok = True
                            break
                    except Exception as _e:
                        print(f"    → Captcha pré-cert tentativa {_t}/2: {type(_e).__name__}")
                    if _t < 2:
                        page.wait_for_timeout(2_000)
                if not _pre_cert_ok:
                    print("    → Captcha pré-cert não resolvido em 2 tentativas. "
                          "Recarregando e refazendo 'Seu certificado digital'...")
                    try:
                        page.reload(wait_until="domcontentloaded", timeout=15_000)
                    except Exception:
                        pass
                    _cert_clicado = False

            page.wait_for_timeout(400)

        if not _cert_clicado:
            print("    [!] 'Seu certificado digital' não encontrado em 40s. Continuando mesmo assim...")

        page.wait_for_timeout(800)

        # Aguarda o avatar aparecer — indica que o OAuth completou.
        # Usa polling porque captcha pode aparecer DURANTE o fluxo OAuth,
        # impedindo o avatar de aparecer dentro de um wait_for simples.
        print("    → Aguardando autenticação no portal...")
        _avatar = page.locator('xpath=//*[@id="avatar-dropdown-trigger"]').first
        _autenticado = False
        _erro_validacao_captcha = False
        _deadline_auth = time.time() + 90  # 90s: OAuth + possível captcha

        while time.time() < _deadline_auth:
            # Avatar apareceu → autenticação concluída
            if _avatar.is_visible():
                _autenticado = True
                break

            # Verifica se o portal rejeitou o captcha durante o OAuth
            try:
                _err_govbr = page.get_by_text("Erro na validação do HCaptcha", exact=False).first
                if _err_govbr.is_visible(timeout=300):
                    print("    [!] 'Erro na validação do HCaptcha' durante autenticação. "
                          "Voltando à página anterior e tentando novamente...")
                    try:
                        page.go_back(wait_until="domcontentloaded", timeout=15_000)
                    except Exception:
                        pass
                    _erro_validacao_captcha = True
                    _fazer_go_back = True
                    break
            except Exception:
                pass

            # Verifica captcha challenge ativo durante o OAuth
            _hc_auth_frames = [
                f for f in page.frames
                if "hcaptcha.com" in (f.url or "")
                and "frame=challenge" in (f.url or "")
            ]
            _captcha_ativo_auth = False
            for _hf in _hc_auth_frames:
                try:
                    _txt = _hf.evaluate("""() => {
                        const c = document.querySelector('.challenge-container');
                        if (!c) return null;
                        const r = c.getBoundingClientRect();
                        if (r.width < 100 || r.height < 100) return null;
                        const p = document.querySelector('.prompt-text');
                        if (!p || !p.textContent.trim()) return null;
                        const b = document.querySelector('.button-submit');
                        if (!b || b.getAttribute('aria-disabled') === 'true') return null;
                        return p.textContent.trim();
                    }""")
                    if _txt:
                        _captcha_ativo_auth = True
                        print(f"    → Captcha durante autenticação: '{_txt}'. Resolvendo...")
                        break
                except Exception:
                    pass

            # Verifica também o checkbox (pode preceder o challenge)
            if not _captcha_ativo_auth:
                try:
                    _captcha_ativo_auth = page.locator(
                        "iframe[src*='hcaptcha.com'][src*='frame=checkbox']"
                    ).is_visible()
                    if _captcha_ativo_auth:
                        print("    → Checkbox hCaptcha detectado durante autenticação. Resolvendo...")
                except Exception:
                    pass

            if _captcha_ativo_auth:
                _oauth_captcha_ok = False
                for _t in range(1, 3):   # máx. 2 tentativas
                    try:
                        if solve_hcaptcha(page):
                            print(f"    [✓] Captcha resolvido (tentativa {_t}/2).")
                            _oauth_captcha_ok = True
                            break
                        print(f"    → Captcha tentativa {_t}/2: solver retornou False.")
                    except Exception as _e:
                        print(f"    → Captcha tentativa {_t}/2: {type(_e).__name__}: {_e}")
                    if _t < 2:
                        page.wait_for_timeout(2_000)
                if not _oauth_captcha_ok:
                    print("    → Captcha OAuth não resolvido em 2 tentativas. "
                          "Voltando e refazendo 'Entrar com gov.br'...")
                    try:
                        page.go_back(wait_until="domcontentloaded", timeout=15_000)
                    except Exception:
                        pass
                    _fazer_go_back = True
                    break  # sai do loop de avatar → outer loop reinicia do gov.br

            page.wait_for_timeout(500)

        if _erro_validacao_captcha:
            continue  # volta para o topo do for: navega de novo para URL_SERVICOS_RF

        if not _autenticado:
            raise Exception(
                "Timeout aguardando autenticação no portal: "
                "avatar não apareceu em 90s após clicar 'Entrar com gov.br'."
            )
        print("    [✓] Autenticado no portal.")
        # Captcha já foi tratado no polling acima (durante o OAuth).
        # Não deve ser solicitado após o avatar aparecer.
        break
    else:
        print(f"    [!] Limite de {MAX_TENTATIVAS_GOVBR} tentativas atingido "
              "para 'Entrar com gov.br'.")


def trocar_perfil_procurador(page, cnpj: str) -> None:
    """Aguarda o intervalo de 30s, depois representa o CNPJ como Procurador
    no portal e navega para a página de pendências.

    Pode ser chamado tanto para o primeiro CNPJ (logo após entrar no portal)
    quanto para os seguintes (sem precisar voltar ao eCAC).
    """
    global _ultimo_troca_cnpj

    # ── Garante intervalo mínimo de 30s entre trocas ──────────────────────────
    _aguardar_intervalo_troca()

    _MAX_TENTATIVAS_REPR = 3

    for _tentativa_repr in range(1, _MAX_TENTATIVAS_REPR + 1):

        if _tentativa_repr > 1:
            print(
                f"    → Retentando representação "
                f"(tentativa {_tentativa_repr}/{_MAX_TENTATIVAS_REPR})..."
            )
            # Fecha dropdown se ainda estiver aberto
            try:
                page.keyboard.press("Escape")
                page.wait_for_timeout(500)
            except Exception:
                pass
            page.wait_for_timeout(2_000)

        # ── Abre menu do avatar ───────────────────────────────────────────────
        print(f"    → Abrindo menu do certificado...")
        avatar = page.locator('xpath=//*[@id="avatar-dropdown-trigger"]').first
        avatar.wait_for(state="visible", timeout=15_000)
        avatar.click()
        page.wait_for_timeout(600)

        # ── Preenche CNPJ ─────────────────────────────────────────────────────
        print(f"    → Digitando CNPJ {cnpj} no campo de representação...")
        campo_cnpj = page.locator('xpath=//*[@id="input-representar-cpfcnpj"]').first
        campo_cnpj.wait_for(state="visible", timeout=10_000)
        campo_cnpj.fill(cnpj)
        page.wait_for_timeout(400)

        # ── Seleciona Procurador ──────────────────────────────────────────────
        print("    → Selecionando 'Procurador' no dropdown...")
        ng_select = page.locator(
            'xpath=//*[@id="formularioRepresentacao"]/form/div/div[2]/br-select/div/div/div[1]/ng-select'
        ).first
        ng_select.wait_for(state="visible", timeout=10_000)
        ng_select.click()
        page.wait_for_timeout(400)

        opcao = page.get_by_role("option", name="Procurador").first
        opcao.wait_for(state="visible", timeout=5_000)
        opcao.click()
        page.wait_for_timeout(400)

        # ── Clica Representar ─────────────────────────────────────────────────
        print("    → Clicando em 'Representar'...")
        btn_representar = page.locator(
            'xpath=//*[@id="formularioRepresentacao"]/form/div/button'
        ).first
        btn_representar.wait_for(state="visible", timeout=10_000)

        # Listener de popup configurado ANTES do clique
        _popups: list = []

        def _on_new_page(p):
            _popups.append(p)

        page.context.on("page", _on_new_page)
        btn_representar.click()

        # Inicia cronômetro imediatamente após clicar em Representar
        _ultimo_troca_cnpj = time.time()

        # ── Aguarda "Carregando" APARECER antes de checar captcha ────────────
        # O captcha pode aparecer POR CIMA do spinner "Carregando".
        # Basta aguardar o spinner aparecer para saber que o servidor recebeu
        # o clique; não esperamos ele sumir — isso ocorre após resolver captcha.
        print("    → Aguardando 'Carregando' aparecer...")
        _carregando = page.locator(
            'xpath=/html/body/app-root/mf-portal-layout/portal-main-layout'
            '/br-loading/div/div/a/div[2]'
        ).first
        _carregando_apareceu = False
        try:
            _carregando.wait_for(state="visible", timeout=8_000)
            _carregando_apareceu = True
            print("    → [Carregando...] detectado.")
        except Exception:
            pass  # spinner não apareceu (resposta muito rápida) — segue

        # ── Espera ativa (até 25 s): erro / popup / captcha / confirmação ─────
        # Prioridade de verificação a cada ~800 ms:
        #   0. Erro "acesso automatizado" (span.mensagemErro) → retentar
        #   1. Popup (nova janela) → captcha em popup
        #   2. iframes hcaptcha VISIVEIS (challenge / checkbox) → captcha inline
        #   3. CNPJ mudou no cabeçalho → representação sem captcha
        _LIMITE_ESPERA_S  = 60
        _deadline_captcha = time.time() + _LIMITE_ESPERA_S
        captcha_tipo      = None    # "popup" | "inline" | None
        _srcs_logados     = False
        _erro_bloqueado   = False

        while time.time() < _deadline_captcha:

            # 0. Mensagem de erro do portal (anti-bot ou falha permanente)
            _err_msg = page.evaluate(
                "() => { const e = document.querySelector('span.mensagemErro'); "
                "return e ? e.textContent.trim() : ''; }"
            )
            if _err_msg:
                _err_lower = _err_msg.lower()
                if "automatizado" in _err_lower or "bloqueado" in _err_lower:
                    print(f"    → [!] Erro anti-bot: '{_err_msg}'")
                    _erro_bloqueado = True
                    break
                if any(p in _err_lower for p in _PALAVRAS_ERRO_PERMANENTE):
                    raise FalhaPermanente(
                        f"Portal recusou CNPJ {cnpj}: '{_err_msg}'"
                    )

            # 1. Popup nova janela
            if _popups:
                captcha_tipo = "popup"
                break

            # 2. Captcha challenge ATIVO — verificado por múltiplos seletores internos
            # ─────────────────────────────────────────────────────────────────────
            # Iframes frame=challenge ficam pré-carregados no DOM mesmo sem captcha
            # ativo. Para evitar falsos positivos, executamos JavaScript DENTRO do
            # iframe (cross-origin acessível pelo Playwright via frame.evaluate) e
            # exigimos que TODOS os critérios abaixo sejam satisfeitos ao mesmo tempo:
            #
            #   1. .challenge-container   → existe E tem dimensões ≥ 100×100 px
            #   2. .prompt-text           → existe E tem texto não-vazio
            #                               (ex: "Toque em todos os seres vivos")
            #   3. .task-grid             → existe (grade de imagens do desafio)
            #   4. .button-submit         → existe E aria-disabled ≠ "true"
            #
            # Se qualquer critério falhar → captcha não está ativo.
            _hc_frames = [
                f for f in page.frames
                if "hcaptcha.com" in (f.url or "")
                and "frame=challenge" in (f.url or "")
            ]

            # Diagnóstico: loga srcs uma vez quando frames aparecerem no DOM
            if _hc_frames and not _srcs_logados:
                _srcs_logados = True
                print(f"    → hcaptcha: {len(_hc_frames)} frame(s) challenge no DOM.")
                for _hf in _hc_frames[:3]:
                    print(f"      src: {(_hf.url or '')[:220]}")

            _captcha_texto = None   # instrução do desafio, se ativo
            for _hf in _hc_frames:
                try:
                    _captcha_texto = _hf.evaluate("""() => {
                        // 1. challenge-container com dimensões reais
                        const container = document.querySelector('.challenge-container');
                        if (!container) return null;
                        const r = container.getBoundingClientRect();
                        if (r.width < 100 || r.height < 100) return null;

                        // 2. prompt-text com instrução preenchida
                        const prompt = document.querySelector('.prompt-text');
                        if (!prompt || !prompt.textContent.trim()) return null;

                        // 3. botão submit habilitado
                        // Nota: NÃO verificamos .task-grid pois só existe no tipo grade 3x3.
                        //       O tipo "imagem única" não tem .task-grid mas é igualmente válido.
                        const btn = document.querySelector('.button-submit');
                        if (!btn || btn.getAttribute('aria-disabled') === 'true') return null;

                        return prompt.textContent.trim();
                    }""")
                    if _captcha_texto:
                        break
                except Exception:
                    pass

            if _captcha_texto:
                print(f"    → Captcha ATIVO: '{_captcha_texto}'. Resolvendo...")
                captcha_tipo = "inline"
                break

            # 3. CNPJ já mudou no cabeçalho? (botão avatar, sempre visível)
            _cnpj_pag = page.evaluate(
                "() => {"
                "  const h = document.querySelector('.ni-pessoa:not(.ni-representante)');"
                "  if (h) return h.textContent.trim();"
                "  const r = document.querySelector('.ni-representacao');"
                "  return r ? r.textContent.trim() : '';"
                "}"
            )
            if re.sub(r"\D", "", _cnpj_pag).zfill(14) == cnpj:
                print("    → Representação concluída sem captcha!")
                break

            # 4. "Carregando" sumiu = servidor respondeu sem captcha aparecer
            if _carregando_apareceu and not _carregando.is_visible():
                print("    → [Carregando...] sumiu. Sem captcha necessário.")
                break

            page.wait_for_timeout(800)

        try:
            page.context.remove_listener("page", _on_new_page)
        except Exception:
            pass

        # ── Erro anti-bot → retentar ou desistir ─────────────────────────────
        if _erro_bloqueado:
            if _tentativa_repr < _MAX_TENTATIVAS_REPR:
                continue  # abre menu, preenche, clica de novo
            raise Exception(
                f"Acesso bloqueado por anti-bot após {_MAX_TENTATIVAS_REPR} "
                "tentativa(s). Reprocessar manualmente mais tarde."
            )

        # ── Resolve captcha conforme tipo ─────────────────────────────────────
        _captcha_repr_ok = True   # False se 2 tentativas falharem → retry repr
        if captcha_tipo == "popup":
            popup = _popups[0]
            print("    → Popup de captcha detectada. Resolvendo...")
            _popup_ok = False
            for tentativa in range(1, 3):   # máx. 2 tentativas
                try:
                    if solve_hcaptcha(popup):
                        print(f"    [✓] Captcha resolvido (tentativa {tentativa}/2).")
                        _popup_ok = True
                        break
                except Exception as e:
                    print(f"    → Captcha tentativa {tentativa}/2: {type(e).__name__}: {e}")
                if tentativa < 2:
                    page.wait_for_timeout(2_000)
            if not _popup_ok:
                _captcha_repr_ok = False
            try:
                popup.wait_for_close(timeout=15_000)
                print("    [✓] Popup do captcha fechada.")
            except Exception:
                pass
        elif captcha_tipo == "inline":
            _inline_ok = False
            for tentativa in range(1, 3):   # máx. 2 tentativas
                try:
                    if solve_hcaptcha(page):
                        print(f"    [✓] Captcha resolvido (tentativa {tentativa}/2).")
                        _inline_ok = True
                        break
                except Exception as e:
                    print(f"    → Captcha tentativa {tentativa}/2: {type(e).__name__}: {e}")
                if tentativa < 2:
                    page.wait_for_timeout(2_000)
            if not _inline_ok:
                _captcha_repr_ok = False
        else:
            print("    → Nenhum captcha detectado. Aguardando confirmação...")

        # Captcha não resolvido em 2 tentativas → reinicia o fluxo de representação
        if not _captcha_repr_ok:
            print("    → Captcha não resolvido em 2 tentativas. Refazendo 'Representar'...")
            if _tentativa_repr < _MAX_TENTATIVAS_REPR:
                continue
            raise Exception(
                f"Captcha não resolvido para CNPJ {cnpj} após "
                f"{_MAX_TENTATIVAS_REPR} tentativa(s)."
            )

        # ── Aguarda "Carregando" DESAPARECER após captcha ────────────────────
        # Quando captcha é resolvido, o "Carregando" ainda está visível por baixo.
        # Esperamos ele sumir para saber que o servidor processou a representação.
        # Se não houve captcha, o "Carregando" já sumiu no loop acima — esta espera
        # retorna imediatamente.
        if captcha_tipo:
            print("    → Aguardando [Carregando...] sumir após captcha...")
        try:
            _carregando.wait_for(state="hidden", timeout=30_000)
            if captcha_tipo:
                print("    → [Carregando...] sumiu.")
        except Exception:
            if captcha_tipo:
                # Página travou no spinner — recarrega e verifica a situação.
                # Se a representação já foi aceita pelo servidor, o CNPJ estará
                # confirmado no cabeçalho após o reload e a automação continua.
                # Se não, a verificação de CNPJ abaixo detecta e retenta.
                print("    → [!] [Carregando...] não sumiu no timeout. Recarregando página...")
                try:
                    page.reload(wait_until="domcontentloaded", timeout=60_000)
                    page.wait_for_timeout(1_500)
                    print(f"    → Página recarregada. URL: {page.url[:80]}. Verificando situação...")
                except Exception as _reload_err:
                    print(f"    → Erro ao recarregar ({type(_reload_err).__name__}): {_reload_err}")

        page.wait_for_timeout(800)

        # ── Verifica erro anti-bot após captcha ───────────────────────────────
        # O portal pode exibir "acesso bloqueado" também DEPOIS de resolver o
        # captcha, quando o servidor processa a representação e rejeita o acesso.
        _err_pos_captcha = page.evaluate(
            "() => { const e = document.querySelector('span.mensagemErro'); "
            "return e ? e.textContent.trim() : ''; }"
        )
        if _err_pos_captcha:
            _epc_lower = _err_pos_captcha.lower()
            if "automatizado" in _epc_lower or "bloqueado" in _epc_lower:
                print(f"    → [!] Erro anti-bot após captcha: '{_err_pos_captcha}'")
                if _tentativa_repr < _MAX_TENTATIVAS_REPR:
                    continue
                raise Exception(
                    f"Acesso bloqueado por anti-bot após captcha — "
                    f"{_MAX_TENTATIVAS_REPR} tentativa(s). Reprocessar manualmente."
                )
            if any(p in _epc_lower for p in _PALAVRAS_ERRO_PERMANENTE):
                raise FalhaPermanente(
                    f"Portal recusou CNPJ {cnpj}: '{_err_pos_captcha}'"
                )

        # ── Verifica que o CNPJ representado realmente mudou ──────────────────
        _cnpj_confirmado = False
        _cnpj_pag_final  = ""
        for _ in range(10):
            _cnpj_pag_final = page.evaluate(
                "() => {"
                "  const h = document.querySelector('.ni-pessoa:not(.ni-representante)');"
                "  if (h) return h.textContent.trim();"
                "  const r = document.querySelector('.ni-representacao');"
                "  return r ? r.textContent.trim() : '';"
                "}"
            )
            if re.sub(r"\D", "", _cnpj_pag_final).zfill(14) == cnpj:
                _cnpj_confirmado = True
                break
            page.wait_for_timeout(500)

        if _cnpj_confirmado:
            break  # ← sucesso, sai do loop de retry

        # CNPJ não confirmado → retentar se ainda houver tentativas
        if _tentativa_repr < _MAX_TENTATIVAS_REPR:
            print(
                f"    → CNPJ não confirmado (portal exibe '{_cnpj_pag_final}'). "
                "Retentando..."
            )
            continue

        raise Exception(
            f"Representação falhou para CNPJ {cnpj} após {_MAX_TENTATIVAS_REPR} "
            f"tentativa(s). Portal ainda exibe: '{_cnpj_pag_final}'."
        )

    print(f"    [✓] Perfil alterado para Procurador do CNPJ {cnpj}.")

    _goto_seguro(page, URL_PENDENCIAS, label="pendências")
    page.wait_for_timeout(500)


def verificar_pendencias(page, cnpj: str, caminho_planilha: str,
                          skip_dctfweb: bool = False,
                          skip_processo: bool = False) -> str:
    """Lê o status de pendências e toma a ação correspondente.

    Parâmetros de controle (usados quando uma das colunas já foi preenchida):
        skip_dctfweb  — True quando coluna D já está preenchida (DCTFWeb concluído).
                        Pula o botão 'Dívida DCTFWeb'; processa apenas Processo Fiscal.
        skip_processo — True quando coluna E já está preenchida (Fiscal concluído).
                        Pula o botão 'Processo Fiscal'; processa apenas DCTFWeb.

    Colunas escritas:
        D — status do DCTFWeb  ('Concluído' / 'Sem débitos' / 'Débitos não compensáveis')
        E — status do Fiscal   ('Concluído' / 'Sem Processos')

    Retorna:
        "sem_pendencia"   — sem débitos
        "nao_compensavel" — débitos não compensáveis
        "concluido"       — ao menos um fluxo foi processado nesta execução
        "desconhecido"    — status inesperado
    """
    print("    → Verificando status de pendências...")
    span_status = page.locator(XPATH_STATUS).first
    span_status.wait_for(state="visible", timeout=30_000)
    texto = (span_status.text_content() or "").strip()
    print(f"    → Status: '{texto}'")

    # ── Sem pendência ─────────────────────────────────────────────────────────
    if texto == "Sem pendência":
        if not skip_dctfweb:
            escrever_coluna_d(caminho_planilha, cnpj, "Sem débitos")
        if not skip_processo:
            escrever_coluna_e(caminho_planilha, cnpj, "Sem Processos")
        return "sem_pendencia"

    # ── Com pendência ─────────────────────────────────────────────────────────
    if texto == "Com pendência":
        btn_dctfweb = page.locator('button[aria-label*="DCTFWeb"]').first

        # Detecta presença dos botões via JS — retorna instantaneamente,
        # sem aguardar timeout de espera (página já carregada após wait_for do span).
        _botoes = page.evaluate(
            """
            () => ({
                dctfweb:  !!document.querySelector('button[aria-label*="DCTFWeb"]'),
                processo: !!document.querySelector(
                    'button[aria-label*="processo fiscal" i]'
                )
            })
            """
        )
        tem_dctfweb  = _botoes["dctfweb"]
        tem_processo = _botoes["processo"]

        # Nenhum botão de ação encontrado
        if not tem_dctfweb and not tem_processo:
            if not skip_dctfweb:
                escrever_coluna_d(caminho_planilha, cnpj, "Débitos não compensáveis")
            if not skip_processo:
                escrever_coluna_e(caminho_planilha, cnpj, "Sem Processos")
            return "nao_compensavel"

        # ── Dívida DCTFWeb ────────────────────────────────────────────────────
        if tem_dctfweb and not skip_dctfweb:
            print("    → Clicando em 'Dívida DCTFWeb'...")
            btn_dctfweb.click()
            print("    [✓] Botão 'Dívida DCTFWeb' clicado.")
            extrair_debitos_dctfweb(page, cnpj, caminho_planilha)
            # extrair_debitos_dctfweb já escreve "Concluído" na coluna D
        elif skip_dctfweb:
            print("    → DCTFWeb já processado anteriormente. Pulando...")

        # ── Processo Fiscal ───────────────────────────────────────────────────
        if tem_processo and not skip_processo:
            _goto_seguro(page, URL_ANALISE_PENDENCIAS, label="análise fiscal")
            page.wait_for_timeout(1_000)
            btn_proc_fresh = page.locator('button[aria-label*="processo fiscal"]').first
            btn_proc_fresh.wait_for(state="visible", timeout=10_000)
            btn_proc_fresh.click()
            print("    [✓] Botão 'Processo Fiscal' clicado.")
            extrair_processo_fiscal(page, cnpj, caminho_planilha)
            # extrair_processo_fiscal já escreve "Concluído" na coluna E
            if not tem_dctfweb and not skip_dctfweb:
                # Sem DCTFWeb e sem skip → marca D como concluído via Fiscal
                escrever_coluna_d(caminho_planilha, cnpj, "Concluído")
        elif tem_processo and skip_processo:
            print("    → Processos Fiscais já processados anteriormente. Pulando...")
        else:
            # Não existe botão de Processo Fiscal
            if not skip_processo:
                escrever_coluna_e(caminho_planilha, cnpj, "Sem Processos")

        return "concluido"

    # ── Status inesperado ─────────────────────────────────────────────────────
    print(f"    [!] Status não reconhecido: '{texto}'")
    return "desconhecido"


# ── Processamento por CNPJ ────────────────────────────────────────────────────

def processar_cnpj(page, cnpj: str, row: pd.Series,
                   caminho_planilha: str, primeiro_do_grupo: bool) -> str:
    """Executa o fluxo completo para um CNPJ no portal de Serviços RF.

    Antes de qualquer navegação, lê as colunas D e E da planilha para
    determinar o que já foi feito:

        D preenchida + E preenchida → linha já concluída, pula tudo.
        D preenchida, E vazia       → faz apenas Processos Fiscais (skip_dctfweb).
        E preenchida, D vazia       → faz apenas Débitos DCTFWeb   (skip_processo).
        Ambas vazias                → executa fluxo completo.

    `primeiro_do_grupo` indica se ainda precisamos navegar para o portal
    (True = primeiro CNPJ do grupo de certificado, faz `ir_para_servicos_rf_e_entrar`).
    """
    print(f"    → Processando CNPJ {cnpj}...")

    # ── Verifica o que já foi processado ─────────────────────────────────────
    val_d, val_e = ler_status_cnpj(caminho_planilha, cnpj)

    if val_d and val_e:
        print(f"    → Já totalmente processado (D='{val_d}', E='{val_e}'). Pulando.")
        return "ja_processado"

    skip_dctfweb  = bool(val_d)   # D preenchida → DCTFWeb já concluído
    skip_processo = bool(val_e)   # E preenchida → Processos Fiscais já concluídos

    if skip_dctfweb:
        print(f"    → Coluna D já preenchida ('{val_d}'). Fará apenas Processos Fiscais.")
    if skip_processo:
        print(f"    → Coluna E já preenchida ('{val_e}'). Fará apenas Débitos DCTFWeb.")

    # ── Navegação e processamento ─────────────────────────────────────────────
    if primeiro_do_grupo:
        ir_para_servicos_rf_e_entrar(page)

    trocar_perfil_procurador(page, cnpj)

    status = verificar_pendencias(page, cnpj, caminho_planilha,
                                   skip_dctfweb=skip_dctfweb,
                                   skip_processo=skip_processo)

    return status


# ── Processamento principal ───────────────────────────────────────────────────

def processar(df: pd.DataFrame, certs: dict[str, tuple[Path, str]],
              caminho_planilha: str) -> None:
    """Itera pela planilha ordenada e realiza o fluxo completo para cada CNPJ.

    Estratégia:
    - Primeiro CNPJ de cada grupo de certificado: login no eCAC + entrada no portal.
    - CNPJs seguintes do mesmo grupo: apenas troca de perfil no portal
      (sem voltar ao eCAC), aguardando o intervalo mínimo de 30s.
    - Troca de certificado: fecha o navegador atual; próximo grupo faz login fresco.
    """
    rows     = list(df.iterrows())
    total    = len(rows)
    col_cnpj = df.columns[0]   # Coluna A = CNPJ
    col_cert = df.columns[2]   # Coluna C = CERTIFICADO

    cert_atual        = None
    browser_aberto    = None   # (p, context, page) | None
    primeiro_do_grupo = True
    _MAX_RETENT_CNPJ    = 2                          # tentativas por CNPJ (inclui a 1ª)
    _retentativas_cnpj: dict[str, int] = {}          # contador por CNPJ

    i = 0
    while i < len(rows):
        idx, row = rows[i]

        cnpj_raw    = re.sub(r"\.0+$", "", str(row[col_cnpj]).strip())
        cnpj        = _normalizar_cnpj(cnpj_raw)
        certificado = str(row[col_cert]).strip()

        # Ignora linhas sem CNPJ válido
        if cnpj in ("", "nan", "None", "00000000000000"):
            print(f"  [{idx + 1}/{total}] CNPJ inválido/vazio. Ignorando linha.")
            i += 1
            continue

        # ── Troca de certificado ──────────────────────────────────────────────
        if certificado != cert_atual:
            print(f"\n{'═' * 60}")
            print(f"  Certificado: {certificado}")
            print(f"{'═' * 60}")

            # Fecha o navegador atual (faz logout no portal antes de fechar)
            if browser_aberto:
                p, context, page_atual = browser_aberto
                _fechar_navegador(p, context, page_atual)
                browser_aberto = None

            chave = _buscar_certificado(certificado, certs)
            if chave is None:
                print(f"  [!] Certificado '{certificado}' não encontrado em {SENHAS_JSON.name}.")
                print(f"       Disponíveis: {', '.join(sorted(certs.keys()))}")
                cert_atual = certificado
                i += 1
                continue

            pfx_path, passphrase = certs[chave]
            atualizar_env_certificado(pfx_path, passphrase)
            cert_atual        = certificado
            primeiro_do_grupo = True

        chave = _buscar_certificado(cert_atual, certs)
        if chave is None:
            i += 1
            continue

        print(f"\n  [{idx + 1}/{total}] CNPJ: {cnpj}")

        # ── Abre browser com certificado (apenas para o primeiro CNPJ do grupo) ─
        if browser_aberto is None:
            try:
                p, context, page = abrir_browser_com_certificado(project_dir=LOGIN_ECAC_DIR)
                browser_aberto    = (p, context, page)
                primeiro_do_grupo = True
                print("    [✓] Browser aberto com certificado digital.")
            except Exception as e:
                print(f"    [!] Erro ao abrir browser ({type(e).__name__}: {e}). Pulando CNPJ {cnpj}.")
                i += 1
                continue

        else:
            p, context, page = browser_aberto

        # ── Processa pendências deste CNPJ ────────────────────────────────────
        _avancar = True
        try:
            status_cnpj = processar_cnpj(page, cnpj, row, caminho_planilha, primeiro_do_grupo)
            # Só marca como "já entrou no portal" se realmente navegou;
            # se a linha foi pulada (ja_processado), o próximo CNPJ ainda
            # precisará fazer ir_para_servicos_rf_e_entrar.
            if status_cnpj != "ja_processado":
                primeiro_do_grupo = False

        except FalhaPermanente as e:
            # Procuração expirada/inválida ou CNPJ não autorizado — não retentar.
            print(f"    [!] Falha permanente — CNPJ {cnpj} ignorado: {e}")
            _fechar_navegador(p, context, page)
            browser_aberto    = None
            primeiro_do_grupo = True

        except Exception as e:
            _retentativas_cnpj[cnpj] = _retentativas_cnpj.get(cnpj, 0) + 1
            _n = _retentativas_cnpj[cnpj]
            print(
                f"    [!] Erro ao processar CNPJ {cnpj} "
                f"(tentativa {_n}/{_MAX_RETENT_CNPJ}): {e}"
            )
            _fechar_navegador(p, context, page)
            browser_aberto    = None
            primeiro_do_grupo = True
            if _n < _MAX_RETENT_CNPJ:
                print(f"    → Reabrindo sessão e retentando CNPJ {cnpj}...")
                _avancar = False   # não incrementa i; próxima iteração refaz login
            else:
                print(
                    f"    [!] Máximo de {_MAX_RETENT_CNPJ} tentativa(s) atingido. "
                    f"Pulando CNPJ {cnpj}."
                )

        if _avancar:
            i += 1

    # ── Fecha o navegador ao terminar ─────────────────────────────────────────
    if browser_aberto:
        p, context, page_final = browser_aberto
        _fechar_navegador(p, context, page_final)


# ── Entrada ───────────────────────────────────────────────────────────────────

def main() -> None:
    # ── Argumentos de linha de comando ────────────────────────────────────────
    parser = argparse.ArgumentParser(description="Automação Débitos em Aberto")
    parser.add_argument(
        "--planilha", metavar="ARQUIVO",
        help="Caminho da planilha .xlsx (pula o seletor gráfico)"
    )
    parser.add_argument(
        "--log", metavar="ARQUIVO",
        help="Grava os logs também em arquivo (ex.: execucao.log)"
    )
    args = parser.parse_args()

    # ── Logging opcional em arquivo ───────────────────────────────────────────
    if args.log:
        logging.basicConfig(
            filename=args.log,
            filemode="a",
            level=logging.DEBUG,
            format="%(asctime)s %(message)s",
            datefmt="%H:%M:%S",
            encoding="utf-8",
        )
        # Redireciona print para também escrever no log
        _log_file = open(args.log, "a", encoding="utf-8", buffering=1)
        class _Tee:
            def __init__(self, *streams): self.streams = streams
            def write(self, data):
                for s in self.streams:
                    s.write(data)
            def flush(self):
                for s in self.streams:
                    s.flush()
        sys.stdout = _Tee(sys.stdout, _log_file)
        sys.stderr = _Tee(sys.stderr, _log_file)

    # Passo 1: seleção da planilha (UI ou argumento)
    if args.planilha:
        planilha = args.planilha
    else:
        planilha = selecionar_planilha()
    if not planilha:
        print("Nenhuma planilha selecionada. Encerrando.")
        sys.exit(0)

    print(f"\nPlanilha: {planilha}")

    # Passo 2: resolve pasta de certificados (padrão → ao lado do script → dialog)
    global CERTIFICADOS_DIR, SENHAS_JSON
    CERTIFICADOS_DIR = _resolver_dir_certificados()
    SENHAS_JSON      = CERTIFICADOS_DIR / "senhas.json"

    # Carrega .env do LoginEcac para que GEMINI_API_KEY esteja em os.environ
    # antes de qualquer chamada ao captcha solver
    _env_path = LOGIN_ECAC_DIR / ".env"
    if _env_path.exists():
        load_dotenv(dotenv_path=_env_path, override=True)
    if not os.environ.get("GEMINI_API_KEY"):
        os.environ["GEMINI_API_KEY"] = _GEMINI_API_KEY_PADRAO

    # Passo 3 (era 2): carrega mapeamento de certificados
    certs = carregar_certificados()

    # Passo 3: leitura e ordenação por certificado (coluna C)
    df           = ler_e_ordenar(planilha)
    col_cert     = df.columns[2]
    certificados = df[col_cert].dropna().unique().tolist()

    print(f"Registros:    {len(df)}")
    print(f"Certificados: {len(certificados)}")
    for cert in certificados:
        qtd    = (df[col_cert] == cert).sum()
        chave  = _buscar_certificado(cert, certs)
        status = "✓" if chave is not None else "✗ NÃO ENCONTRADO"
        print(f"  {status}  {cert}  ({qtd} CNPJ{'s' if qtd > 1 else ''})")

    # Passo 4: fluxo completo para cada CNPJ
    processar(df, certs, planilha)

    print("\nProcessamento concluído.")


if __name__ == "__main__":
    main()
